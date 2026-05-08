# tests/test_exporter.py
import io
import openpyxl
from app.exporter import build_excel, build_word
from app.models import SessionData, Student, Entry
from app.seating import compute_seating
import uuid


def _make_session() -> SessionData:
    student = Student(id="s1", last_name="Müller", first_name="Anna", class_name="10a")
    entry = Entry(
        id=str(uuid.uuid4()),
        student_id="s1",
        subject="Mathematik",
        duration_minutes=45,
        aids="Taschenrechner",
        teacher="Fr. Schmidt",
        room="A",
        desk=1,
        seat=1,
    )
    return SessionData(students=[student], entries=[entry])


def test_excel_has_three_sheets():
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    assert set(wb.sheetnames) == {"Raum 1", "Raum 2", "Raum 3"}


def test_excel_grid_has_title_and_lehrpult_rows():
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    # Row 1 = title (merged A1:H1), Row 2 = Lehrpult (merged A2:H2)
    assert "Raum 1" in (ws.cell(1, 1).value or "")
    assert (ws.cell(2, 1).value or "").strip() == "Lehrpult"


def test_excel_grid_seat_1_1_holds_first_entry():
    # Student at desk=1 seat=1 — desk 1 ist nun unten rechts (row 7, col 7)
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    text = str(ws.cell(7, 7).value or "")
    assert "Müller, Anna" in text
    assert "10a" in text
    assert "Mathematik" in text
    assert "45 min" in text
    assert "Fr. Schmidt" in text
    assert "Taschenrechner" in text


def test_excel_grid_empty_seat_shows_only_label():
    # Desk 1, Seat 2 → row 7, col 8
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    text = str(ws.cell(7, 8).value or "").strip()
    assert text == "T1.S2"


def test_excel_grid_second_desk_first_seat_is_empty_with_label():
    # Desk 2 ist unten zweite-von-rechts → row 7, col 5
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    text = str(ws.cell(7, 5).value or "").strip()
    assert text == "T2.S1"


def test_excel_grid_top_row_corresponds_to_desks_13_to_16():
    # Top row (row 4) hält von links nach rechts: T16 T15 T14 T13
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    assert str(ws.cell(4, 1).value or "").strip() == "T16.S1"
    assert str(ws.cell(4, 2).value or "").strip() == "T16.S2"
    assert str(ws.cell(4, 7).value or "").strip() == "T13.S1"
    assert str(ws.cell(4, 8).value or "").strip() == "T13.S2"


def test_excel_grid_bottom_left_desk_is_t4():
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    ws = wb["Raum 1"]
    assert str(ws.cell(7, 1).value or "").strip() == "T4.S1"
    assert str(ws.cell(7, 2).value or "").strip() == "T4.S2"


def test_excel_grid_landscape():
    session = _make_session()
    plan = compute_seating(session)
    buf = build_excel(plan)
    wb = openpyxl.load_workbook(io.BytesIO(buf))
    for name in ("Raum 1", "Raum 2", "Raum 3"):
        assert wb[name].page_setup.orientation == "landscape"


def test_word_returns_bytes():
    session = _make_session()
    plan = compute_seating(session)
    buf = build_word(plan)
    assert isinstance(buf, bytes)
    assert len(buf) > 100  # non-empty DOCX
